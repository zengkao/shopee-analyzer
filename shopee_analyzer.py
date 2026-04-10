"""
蝦皮訂單 & 廣告費分析工具
讓蝦皮賣家可以用自己的數據，分析營業額、廣告費、優惠券、平台抽成的趨勢。
"""

import os
import sys
import io
import re
import threading
from datetime import datetime

try:
    import tkinter as tk
    from tkinter import filedialog, scrolledtext, messagebox, ttk
except ImportError:
    print("錯誤：缺少 tkinter")
    sys.exit(1)

try:
    import pandas as pd
except ImportError:
    messagebox.showerror("錯誤", "缺少 pandas 套件，請執行 pip install pandas")
    sys.exit(1)

try:
    import openpyxl
except ImportError:
    messagebox.showerror("錯誤", "缺少 openpyxl 套件，請執行 pip install openpyxl")
    sys.exit(1)

try:
    import msoffcrypto
except ImportError:
    msoffcrypto = None


# ── 核心分析邏輯 ──────────────────────────────────────────

def open_encrypted_xlsx(path, password):
    if msoffcrypto is None:
        raise RuntimeError("檔案有密碼保護，但缺少 msoffcrypto 套件。\n請執行：pip install msoffcrypto-tool")
    with open(path, "rb") as f:
        file = msoffcrypto.OfficeFile(f)
        file.load_key(password=password)
        decrypted = io.BytesIO()
        file.decrypt(decrypted)
        decrypted.seek(0)
        return decrypted


def load_order_file(path, password=None):
    try:
        wb = openpyxl.load_workbook(path, read_only=False)
    except Exception:
        if password:
            dec = open_encrypted_xlsx(path, password)
            wb = openpyxl.load_workbook(dec, read_only=False)
        else:
            return None

    ws = wb.active
    headers = [ws.cell(row=1, column=i).value for i in range(1, ws.max_column + 1)]
    data = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data.append(row)
    wb.close()
    return pd.DataFrame(data, columns=headers)


def load_ad_file(path):
    for enc in ['utf-8-sig', 'utf-8', 'big5', 'cp950']:
        try:
            df = pd.read_csv(path, skiprows=7, encoding=enc)
            if '花費' in df.columns:
                return df
        except Exception:
            continue
    return None


def to_num(series):
    return pd.to_numeric(series.astype(str).str.replace(',', '').str.replace('%', ''), errors='coerce')


def extract_month_label(filename):
    m = re.search(r'(\d{4})(\d{2})\d{2}_\d{8}', filename)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        return f"{year}/{month}月", year * 100 + month

    m = re.search(r'(\d{4})_(\d{2})_\d{2}-\d{4}_\d{2}_\d{2}', filename)
    if m:
        year, month = int(m.group(1)), int(m.group(2))
        return f"{year}/{month}月", year * 100 + month

    return os.path.basename(filename), 0


def analyze_orders(order_files, password=None, log=None):
    results = {}
    for path in order_files:
        label, sort_key = extract_month_label(os.path.basename(path))
        if log:
            log(f"  讀取訂單：{os.path.basename(path)} → {label}")

        df = load_order_file(path, password)
        if df is None:
            if log:
                log(f"  ⚠ 無法開啟 {os.path.basename(path)}，可能需要密碼")
            continue

        required = ['訂單編號', '商品總價', '買家總支付金額', '優惠代碼',
                     '蝦皮負擔優惠券', '蝦幣折抵', '成交手續費', '其他服務費', '金流與系統處理費']
        missing = [c for c in required if c not in df.columns]
        if missing:
            if log:
                log(f"  ⚠ 缺少欄位 {missing}，跳過")
            continue

        u = df.drop_duplicates(subset='訂單編號', keep='first')
        revenue = pd.to_numeric(u['商品總價'], errors='coerce').sum()
        buyer_paid = pd.to_numeric(u['買家總支付金額'], errors='coerce').sum()

        has_coupon = u[u['優惠代碼'].notna() & (u['優惠代碼'] != '')]

        results[label] = {
            'sort_key': sort_key,
            'orders': len(u),
            'coupon_orders': len(has_coupon),
            'revenue': revenue,
            'buyer_paid': buyer_paid,
            'total_discount': revenue - buyer_paid,
            'shopee_coupon': pd.to_numeric(u['蝦皮負擔優惠券'], errors='coerce').sum(),
            'coin_discount': pd.to_numeric(u['蝦幣折抵'], errors='coerce').sum(),
            'tx_fee': pd.to_numeric(u['成交手續費'], errors='coerce').sum(),
            'other_fee': pd.to_numeric(u['其他服務費'], errors='coerce').sum(),
            'payment_fee': pd.to_numeric(u['金流與系統處理費'], errors='coerce').sum(),
            'platform_total': (pd.to_numeric(u['成交手續費'], errors='coerce').sum() +
                               pd.to_numeric(u['其他服務費'], errors='coerce').sum() +
                               pd.to_numeric(u['金流與系統處理費'], errors='coerce').sum()),
        }
    return dict(sorted(results.items(), key=lambda x: x[1]['sort_key']))


def analyze_ads(ad_files, log=None):
    results = {}
    for path in ad_files:
        label, sort_key = extract_month_label(os.path.basename(path))
        if log:
            log(f"  讀取廣告：{os.path.basename(path)} → {label}")

        df = load_ad_file(path)
        if df is None:
            if log:
                log(f"  ⚠ 無法讀取 {os.path.basename(path)}")
            continue

        ad_spend = to_num(df['花費']).sum()
        clicks = to_num(df['點擊數']).sum()
        conversions = to_num(df['轉換數']).sum()

        results[label] = {
            'sort_key': sort_key,
            'ad_spend': ad_spend,
            'ad_sales': to_num(df['銷售金額']).sum(),
            'clicks': clicks,
            'impressions': to_num(df['瀏覽數']).sum(),
            'conversions': conversions,
            'roas': to_num(df['銷售金額']).sum() / ad_spend if ad_spend > 0 else 0,
            'cpc': ad_spend / clicks if clicks > 0 else 0,
            'conv_rate': conversions / clicks * 100 if clicks > 0 else 0,
        }
    return dict(sorted(results.items(), key=lambda x: x[1]['sort_key']))


def generate_report(order_results, ad_results):
    lines = []
    W = 12

    all_months = list(dict.fromkeys(list(order_results.keys()) + list(ad_results.keys())))
    header = f"{'':30s}" + "".join(f"{m:>{W}s}" for m in all_months)
    sep = "─" * (30 + W * len(all_months))

    lines.append("=" * len(sep))
    lines.append("【蝦皮賣場分析報告】")
    lines.append(f"產生時間：{datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")
    lines.append("=" * len(sep))

    def fmtrow(label, vals):
        lines.append(f"  {label:28s}{''.join(vals)}")

    # ── 訂單 ──
    if order_results:
        lines.append(f"\n{header}")
        lines.append(sep)

        def oval(key, fmt=',.0f', prefix='$'):
            vals = []
            for m in all_months:
                if m in order_results:
                    v = order_results[m][key]
                    if prefix == '$':
                        vals.append(f"${v:>{W-1}{fmt}}")
                    else:
                        vals.append(f"{v:>{W}{fmt}}")
                else:
                    vals.append(f"{'N/A':>{W}s}")
            return vals

        fmtrow('訂單數', oval('orders', 'd', ''))
        fmtrow('有優惠券訂單數', oval('coupon_orders', 'd', ''))

        rate_vals = []
        for m in all_months:
            if m in order_results:
                r = order_results[m]
                pct = r['coupon_orders'] / r['orders'] * 100 if r['orders'] > 0 else 0
                rate_vals.append(f"{pct:>{W-1}.1f}%")
            else:
                rate_vals.append(f"{'N/A':>{W}s}")
        fmtrow('優惠券使用率', rate_vals)

        lines.append(sep)
        fmtrow('商品總價（營業額）', oval('revenue'))
        fmtrow('買家總支付金額', oval('buyer_paid'))
        fmtrow('所有折扣合計', oval('total_discount'))
        fmtrow('  └ 蝦皮負擔優惠券', oval('shopee_coupon'))
        fmtrow('  └ 蝦幣折抵', oval('coin_discount'))

        lines.append(sep)
        fmtrow('成交手續費', oval('tx_fee'))
        fmtrow('其他服務費', oval('other_fee'))
        fmtrow('金流處理費', oval('payment_fee'))
        fmtrow('平台費用小計', oval('platform_total'))

    # ── 廣告 ──
    if ad_results:
        lines.append(sep)
        vals = []
        for m in all_months:
            if m in ad_results:
                vals.append(f"${ad_results[m]['ad_spend']:>{W-1},.0f}")
            else:
                vals.append(f"{'N/A':>{W}s}")
        fmtrow('廣告費', vals)

    # ── 總成本 ──
    if order_results and ad_results:
        lines.append(sep)
        for row_label, calc in [
            ('總成本(平台+廣告)', lambda o, a: o['platform_total'] + a['ad_spend']),
            ('實際到手估算', lambda o, a: o['buyer_paid'] - o['platform_total'] - a['ad_spend']),
        ]:
            vals = []
            for m in all_months:
                if m in order_results and m in ad_results:
                    v = calc(order_results[m], ad_results[m])
                    vals.append(f"${v:>{W-1},.0f}")
                else:
                    vals.append(f"{'N/A':>{W}s}")
            fmtrow(row_label, vals)

    # ── 費率 ──
    lines.append(f"\n{sep}")
    lines.append(f"  {'【費率分析】'}")
    lines.append(sep)

    if order_results:
        vals = []
        for m in all_months:
            if m in order_results:
                o = order_results[m]
                r = o['platform_total'] / o['revenue'] * 100 if o['revenue'] > 0 else 0
                vals.append(f"{r:>{W-1}.1f}%")
            else:
                vals.append(f"{'N/A':>{W}s}")
        fmtrow('平台費率', vals)

    if ad_results and order_results:
        vals = []
        for m in all_months:
            if m in order_results and m in ad_results:
                r = ad_results[m]['ad_spend'] / order_results[m]['revenue'] * 100 if order_results[m]['revenue'] > 0 else 0
                vals.append(f"{r:>{W-1}.1f}%")
            else:
                vals.append(f"{'N/A':>{W}s}")
        fmtrow('廣告費率(廣告/營收)', vals)

        vals = []
        for m in all_months:
            if m in order_results and m in ad_results:
                o = order_results[m]
                t = (o['platform_total'] + ad_results[m]['ad_spend']) / o['revenue'] * 100 if o['revenue'] > 0 else 0
                vals.append(f"{t:>{W-1}.1f}%")
            else:
                vals.append(f"{'N/A':>{W}s}")
        fmtrow('總成本率', vals)

    # ── 廣告效率 ──
    if ad_results:
        lines.append(f"\n{sep}")
        lines.append(f"  {'【廣告效率】'}")
        lines.append(sep)

        for label, key, fmt, prefix in [
            ('ROAS', 'roas', '.1f', ''),
            ('CPC(每次點擊成本)', 'cpc', '.2f', '$'),
            ('廣告轉換率', 'conv_rate', '.1f', '%'),
        ]:
            vals = []
            for m in all_months:
                if m in ad_results:
                    v = ad_results[m][key]
                    if prefix == '$':
                        vals.append(f"${v:>{W-1}{fmt}}")
                    elif prefix == '%':
                        vals.append(f"{v:>{W-1}{fmt}}%")
                    else:
                        vals.append(f"{v:>{W}{fmt}}")
                else:
                    vals.append(f"{'N/A':>{W}s}")
            fmtrow(label, vals)

    # ── 月度變化 ──
    months_with_both = [m for m in all_months if m in order_results and m in ad_results]
    if len(months_with_both) >= 2:
        base = months_with_both[0]
        bo = order_results[base]
        ba = ad_results[base]

        lines.append(f"\n{sep}")
        lines.append(f"  【月度變化】以 {base} 為基準")
        lines.append(sep)

        for label, get_val in [
            ('營業額變化', lambda m: order_results[m]['revenue'] / bo['revenue'] * 100 - 100),
            ('訂單數變化', lambda m: order_results[m]['orders'] / bo['orders'] * 100 - 100),
            ('廣告費變化', lambda m: ad_results[m]['ad_spend'] / ba['ad_spend'] * 100 - 100),
        ]:
            vals = []
            for m in months_with_both:
                if m == base:
                    vals.append(f"{'基準':>{W}s}")
                else:
                    vals.append(f"{get_val(m):>{W-1}+.1f}%")
            fmtrow(label, vals)

        vals = []
        base_net = bo['buyer_paid'] - bo['platform_total'] - ba['ad_spend']
        for m in months_with_both:
            if m == base:
                vals.append(f"{'基準':>{W}s}")
            else:
                cur_o = order_results[m]
                cur_net = cur_o['buyer_paid'] - cur_o['platform_total'] - ad_results[m]['ad_spend']
                vals.append(f"{(cur_net / base_net * 100 - 100):>{W-1}+.1f}%")
        fmtrow('實際到手變化', vals)

    lines.append("\n" + "=" * len(sep))
    lines.append("報告生成完畢")
    lines.append("=" * len(sep))

    return "\n".join(lines)


# ── GUI ──────────────────────────────────────────────────

class ShopeeAnalyzerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("蝦皮訂單 & 廣告費分析工具")
        self.root.geometry("820x700")
        self.root.resizable(True, True)

        self.order_files = []
        self.ad_files = []

        self._build_ui()

    def _build_ui(self):
        # ── 標題 ──
        title_frame = tk.Frame(self.root, bg="#EE4D2D", pady=8)
        title_frame.pack(fill=tk.X)
        tk.Label(title_frame, text="蝦皮訂單 & 廣告費分析工具",
                 font=("Microsoft JhengHei", 16, "bold"),
                 fg="white", bg="#EE4D2D").pack()
        tk.Label(title_frame, text="讓數據說話，拒絕黑箱",
                 font=("Microsoft JhengHei", 10),
                 fg="white", bg="#EE4D2D").pack()

        # ── 輸入區 ──
        input_frame = ttk.LabelFrame(self.root, text="資料輸入", padding=10)
        input_frame.pack(fill=tk.X, padx=10, pady=(10, 5))

        # 1. 訂單報表
        row = 0
        ttk.Label(input_frame, text="① 訂單報表 (.xlsx)：",
                  font=("Microsoft JhengHei", 10, "bold")).grid(row=row, column=0, sticky=tk.W, pady=(0, 2))
        row += 1
        self.order_listvar = tk.StringVar(value="尚未選擇檔案")
        self.order_label = ttk.Label(input_frame, textvariable=self.order_listvar,
                                      foreground="gray", wraplength=550)
        self.order_label.grid(row=row, column=0, sticky=tk.W, padx=(10, 0))
        btn_frame1 = tk.Frame(input_frame)
        btn_frame1.grid(row=row, column=1, padx=(10, 0))
        ttk.Button(btn_frame1, text="選擇檔案", command=self._pick_order_files).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame1, text="選擇資料夾", command=self._pick_order_folder).pack(side=tk.LEFT, padx=2)

        # 2. 廣告報表
        row += 1
        ttk.Separator(input_frame, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=8)
        row += 1
        ttk.Label(input_frame, text="② 廣告報表 (.csv)：",
                  font=("Microsoft JhengHei", 10, "bold")).grid(row=row, column=0, sticky=tk.W, pady=(0, 2))
        row += 1
        self.ad_listvar = tk.StringVar(value="尚未選擇檔案")
        self.ad_label = ttk.Label(input_frame, textvariable=self.ad_listvar,
                                   foreground="gray", wraplength=550)
        self.ad_label.grid(row=row, column=0, sticky=tk.W, padx=(10, 0))
        btn_frame2 = tk.Frame(input_frame)
        btn_frame2.grid(row=row, column=1, padx=(10, 0))
        ttk.Button(btn_frame2, text="選擇檔案", command=self._pick_ad_files).pack(side=tk.LEFT, padx=2)
        ttk.Button(btn_frame2, text="選擇資料夾", command=self._pick_ad_folder).pack(side=tk.LEFT, padx=2)

        # 3. 密碼
        row += 1
        ttk.Separator(input_frame, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=8)
        row += 1
        ttk.Label(input_frame, text="③ 訂單檔案密碼（無密碼留空）：",
                  font=("Microsoft JhengHei", 10, "bold")).grid(row=row, column=0, sticky=tk.W, pady=(0, 2))
        row += 1
        self.password_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.password_var, show="*", width=30).grid(
            row=row, column=0, sticky=tk.W, padx=(10, 0))

        # 4. 輸出路徑
        row += 1
        ttk.Separator(input_frame, orient=tk.HORIZONTAL).grid(row=row, column=0, columnspan=2, sticky=tk.EW, pady=8)
        row += 1
        ttk.Label(input_frame, text="④ 報告儲存位置（選填）：",
                  font=("Microsoft JhengHei", 10, "bold")).grid(row=row, column=0, sticky=tk.W, pady=(0, 2))
        row += 1
        self.output_var = tk.StringVar()
        ttk.Entry(input_frame, textvariable=self.output_var, width=60).grid(
            row=row, column=0, sticky=tk.W, padx=(10, 0))
        ttk.Button(input_frame, text="瀏覽", command=self._pick_output).grid(
            row=row, column=1, padx=(10, 0))

        input_frame.columnconfigure(0, weight=1)

        # ── 執行按鈕 ──
        btn_frame = tk.Frame(self.root, pady=5)
        btn_frame.pack(fill=tk.X, padx=10)
        self.run_btn = tk.Button(btn_frame, text="開始分析", font=("Microsoft JhengHei", 12, "bold"),
                                  bg="#EE4D2D", fg="white", padx=30, pady=5,
                                  command=self._run_analysis)
        self.run_btn.pack()

        # ── 結果區 ──
        result_frame = ttk.LabelFrame(self.root, text="分析結果", padding=5)
        result_frame.pack(fill=tk.BOTH, expand=True, padx=10, pady=(5, 10))

        self.result_text = scrolledtext.ScrolledText(
            result_frame, wrap=tk.NONE, font=("Consolas", 9),
            state=tk.DISABLED)
        self.result_text.pack(fill=tk.BOTH, expand=True)

        # 水平捲軸
        h_scroll = ttk.Scrollbar(result_frame, orient=tk.HORIZONTAL, command=self.result_text.xview)
        h_scroll.pack(fill=tk.X)
        self.result_text.configure(xscrollcommand=h_scroll.set)

    # ── 檔案選擇 ──

    def _pick_order_files(self):
        files = filedialog.askopenfilenames(
            title="選擇訂單報表",
            filetypes=[("Excel 檔案", "*.xlsx"), ("所有檔案", "*.*")])
        if files:
            self.order_files = list(files)
            names = [os.path.basename(f) for f in files]
            self.order_listvar.set(f"已選擇 {len(files)} 個檔案：{', '.join(names)}")
            self.order_label.configure(foreground="black")

    def _pick_order_folder(self):
        folder = filedialog.askdirectory(title="選擇訂單報表資料夾")
        if folder:
            import glob
            files = sorted(glob.glob(os.path.join(folder, "*.xlsx")))
            if files:
                self.order_files = files
                self.order_listvar.set(f"資料夾：{folder}（{len(files)} 個 xlsx 檔）")
                self.order_label.configure(foreground="black")
            else:
                messagebox.showwarning("提示", f"在 {folder} 找不到 .xlsx 檔案")

    def _pick_ad_files(self):
        files = filedialog.askopenfilenames(
            title="選擇廣告報表",
            filetypes=[("CSV 檔案", "*.csv"), ("所有檔案", "*.*")])
        if files:
            self.ad_files = list(files)
            names = [os.path.basename(f) for f in files]
            self.ad_listvar.set(f"已選擇 {len(files)} 個檔案：{', '.join(names)}")
            self.ad_label.configure(foreground="black")

    def _pick_ad_folder(self):
        folder = filedialog.askdirectory(title="選擇廣告報表資料夾")
        if folder:
            import glob
            files = sorted(glob.glob(os.path.join(folder, "*.csv")))
            if files:
                self.ad_files = files
                self.ad_listvar.set(f"資料夾：{folder}（{len(files)} 個 csv 檔）")
                self.ad_label.configure(foreground="black")
            else:
                messagebox.showwarning("提示", f"在 {folder} 找不到 .csv 檔案")

    def _pick_output(self):
        path = filedialog.asksaveasfilename(
            title="儲存報告",
            defaultextension=".txt",
            filetypes=[("文字檔", "*.txt")],
            initialfile=f"分析報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
        if path:
            self.output_var.set(path)

    # ── 分析 ──

    def _log(self, msg):
        self.result_text.configure(state=tk.NORMAL)
        self.result_text.insert(tk.END, msg + "\n")
        self.result_text.see(tk.END)
        self.result_text.configure(state=tk.DISABLED)
        self.root.update_idletasks()

    def _run_analysis(self):
        if not self.order_files and not self.ad_files:
            messagebox.showwarning("提示", "請至少選擇訂單報表或廣告報表")
            return

        self.run_btn.configure(state=tk.DISABLED, text="分析中...")
        self.result_text.configure(state=tk.NORMAL)
        self.result_text.delete("1.0", tk.END)
        self.result_text.configure(state=tk.DISABLED)

        threading.Thread(target=self._do_analysis, daemon=True).start()

    def _do_analysis(self):
        try:
            password = self.password_var.get().strip() or None

            order_results = {}
            ad_results = {}

            if self.order_files:
                self._log("正在分析訂單資料...")
                order_results = analyze_orders(self.order_files, password, log=self._log)
                self._log(f"  完成，共 {len(order_results)} 個月份\n")

            if self.ad_files:
                self._log("正在分析廣告資料...")
                ad_results = analyze_ads(self.ad_files, log=self._log)
                self._log(f"  完成，共 {len(ad_results)} 個月份\n")

            if not order_results and not ad_results:
                self._log("⚠ 沒有成功讀取任何資料，請檢查檔案格式或密碼是否正確。")
                return

            report = generate_report(order_results, ad_results)
            self._log(report)

            # 儲存
            output_path = self.output_var.get().strip()
            if output_path:
                with open(output_path, 'w', encoding='utf-8') as f:
                    f.write(report)
                self._log(f"\n報告已儲存至：{output_path}")
            else:
                # 自動存到第一個檔案的同目錄
                base_dir = os.path.dirname(self.order_files[0] if self.order_files else self.ad_files[0])
                auto_path = os.path.join(base_dir, f"分析報告_{datetime.now().strftime('%Y%m%d_%H%M%S')}.txt")
                with open(auto_path, 'w', encoding='utf-8') as f:
                    f.write(report)
                self._log(f"\n報告已自動儲存至：{auto_path}")

        except Exception as e:
            self._log(f"\n錯誤：{e}")
        finally:
            self.root.after(0, lambda: self.run_btn.configure(state=tk.NORMAL, text="開始分析"))


def main():
    root = tk.Tk()
    app = ShopeeAnalyzerApp(root)
    root.mainloop()


if __name__ == "__main__":
    main()
